### OJK DYNAMIC WEB SCRAPING USING SELENIUM ###

import re
import time
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── 1. HELPER FUNCTIONS
def teks_td(td) -> str:
    div = td.find('div')
    raw = div.get_text(separator=' ', strip=True) if div else td.get_text(separator=' ', strip=True)
    return re.sub(r'\s+', ' ', raw).strip() or "0"


def ambil_jumlah_dari_label_row(info: dict) -> str:
    tds = info['tds']
    label_col = info['label_col']
    n = len(tds)

    if n >= 8:
        return teks_td(tds[7])
    if n == 2:
        container = tds[1 - label_col] if label_col == 0 else tds[0]
        inner_table = container.find('table')
        if inner_table:
            inner_row = inner_table.find('tr')
            if inner_row:
                inner_tds = inner_row.find_all('td')
                if len(inner_tds) >= 6:
                    return teks_td(inner_tds[5])

    label_td = tds[label_col]
    label_table = label_td.find_parent('table')
    if label_table:
        outer_td = label_table.find_parent('td')
        if outer_td:
            outer_tr = outer_td.find_parent('tr')
            if outer_tr:
                outer_tds = outer_tr.find_all('td', recursive=False)
                for o_td in outer_tds:
                    if o_td == outer_td: continue
                    data_table = o_td.find('table')
                    if data_table:
                        label_trs = label_table.find_all('tr')
                        tr_label = label_td.find_parent('tr')
                        try:
                            row_idx = label_trs.index(tr_label)
                        except ValueError:
                            row_idx = 0
                        data_trs = data_table.find_all('tr')
                        if row_idx < len(data_trs):
                            data_tds = data_trs[row_idx].find_all('td')
                            if len(data_tds) >= 6:
                                return teks_td(data_tds[5])
    return "Tidak Ditemukan"


def cari_html_laporan(driver, max_depth=3):
    def cek_dan_cari(depth):
        soup_cek = BeautifulSoup(driver.page_source, 'html.parser')
        baris_cek = [tr for tr in soup_cek.find_all('tr') if len(tr.find_all('td', recursive=False)) >= 7]
        if baris_cek: return driver.page_source
        if depth >= max_depth: return None
        iframes = driver.find_elements(By.TAG_NAME, 'iframe')
        for idx, frm in enumerate(iframes):
            try:
                driver.switch_to.frame(frm)
                hasil = cek_dan_cari(depth + 1)
                if hasil: return hasil
                driver.switch_to.parent_frame()
            except Exception:
                try:
                    driver.switch_to.parent_frame()
                except Exception:
                    pass
        return None

    return cek_dan_cari(0)


# ── 2. SETUP CHROME & VARIABEL RISET ──────────────────────────────────────────
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=chrome_options)

website = 'https://ojk.go.id/id/kanal/perbankan/data-dan-statistik/laporan-keuangan-perbankan/default.aspx'

# Setup Variabel Panel Data
target_tahun = "2020"
daftar_bank = ["112-PT BPD DAERAH ISTIMEWA YOGYAKARTA", "113-PT BPD JAWA TENGAH", "114-PT BPD JAWA TIMUR Tbk", "115-PT BPD JAMBI ", "116-PT BANK ACEH", "117-PT BPD SUMATERA UTARA ",
               "118-PT BANK NAGARI", "119-PT BPD RIAU DAN KEPULAUAN RIAU ", "120-PT BPD SUMATERA SELATAN DAN BANGKA BELITUNG", "121-PT BPD LAMPUNG", "122-PT BPD KALIMANTAN SELATAN",
               "123-BPD KALIMANTAN BARAT", "124-PT BPD KALIMANTAN TIMUR DAN KALIMANTAN UTARA", "125-PT BPD KALIMANTAN TENGAH", "126-PT BPD SULAWESI SELATAN DAN SULAWESI BARAT ", "127-PT BPD SULAWESI UTARA GORONTALO",
               "128-PT BPD NUSA TENGGARA BARAT", "129-PT BPD BALI ", "130-PT BPD NUSA TENGGARA TIMUR ", "131-PT BPD MALUKU DAN MALUKU UTARA", "132-PT BPD PAPUA", "133-PT BPD BENGKULU", "134-PT.  BPD SULAWESI TENGAH", "135-PT BPD SULAWESI TENGGARA"]
nama_laporan = "Kualitas Aktiva Produktif"
daftar_bulan = ["Maret", "Juni"]

# Tempat penampungan data Excel
data_rows_excel = []
nomor_urut = 1
PATTERN_KREDIT = re.compile(r'7\.\s+Kredit\s*$')
PATTERN_UMKM = re.compile(r'a\.\s+Debitur\s+Usaha\s+Mikro,\s+Kecil\s+dan\s+Menengah\s+\(UMKM\)\s*$')

# ── 3. LOOPING KUARTALAN ──────────────────────────────────────────────────────
for target_bank in daftar_bank:
    print(f"\n{'='*80}")
    print(f" MEMULAI EKSTRAKSI UNTUK BANK: {target_bank}")
    print(f"{'='*80}")

    for target_bulan in daftar_bulan:
        print(f"\n--- Mengambil Bulan: {target_bulan.upper()} {target_tahun} ---")

        # Reload website setiap iterasi (Mencegah error ExtJS)
        driver.get(website)
        wait = WebDriverWait(driver, 20)

        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'devframe')))
        wait.until(EC.frame_to_be_available_and_switch_to_it((By.ID, 'MainReportFrame')))

        # Form: All Matches
        all_matches_button = wait.until(EC.element_to_be_clickable((By.ID, 'R-boxLabelEl')))
        driver.execute_script("arguments[0].click();", all_matches_button)

        # Form: Tahun
        kotak_tahun = wait.until(EC.element_to_be_clickable((By.ID, 'Year-inputEl')))
        kotak_tahun.click()
        time.sleep(1)
        kotak_tahun.send_keys(Keys.COMMAND + 'a')
        time.sleep(1)
        kotak_tahun.send_keys(Keys.BACKSPACE)
        kotak_tahun.send_keys(target_tahun)
        time.sleep(1)
        kotak_tahun.send_keys(Keys.ENTER)

        # Form: Bulan
        panah_bulan = wait.until(EC.element_to_be_clickable((By.ID, 'ext-gen1060')))
        driver.execute_script("arguments[0].click();", panah_bulan)
        xpath_bulan = f"//li[contains(@class, 'x-boundlist-item') and text()='{target_bulan}']"
        pilihan_bulan = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_bulan)))
        driver.execute_script("arguments[0].click();", pilihan_bulan)

        # Form: Bank
        panah_bank = wait.until(EC.element_to_be_clickable((By.ID, 'ext-gen1069')))
        driver.execute_script("arguments[0].click();", panah_bank)
        kotak_bank = wait.until(EC.element_to_be_clickable((By.ID, 'BankCodeSearchField-inputEl')))
        kotak_bank.click()
        time.sleep(1)
        kotak_bank.send_keys(target_bank)
        xpath_bank = f"//span[contains(@class, 'x-tree-node-text') and text()='{target_bank}']"
        pilihan_bank = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_bank)))
        driver.execute_script("arguments[0].click();", pilihan_bank)

        # Form: Laporan
        xpath_laporan = f"//tr[.//span[contains(text(), '{nama_laporan}')]]//input[contains(@class, 'x-tree-checkbox')]"
        checkbox_laporan = wait.until(EC.element_to_be_clickable((By.XPATH, xpath_laporan)))
        driver.execute_script("arguments[0].click();", checkbox_laporan)

        # Tampilkan Data
        btn_show = wait.until(EC.element_to_be_clickable((By.ID, 'ShowReportButton-btnIconEl')))
        driver.execute_script("arguments[0].click();", btn_show)
        print(f"Menunggu {target_bulan} dimuat...")
        time.sleep(10)

        # Scraping HTML
        html_laporan = cari_html_laporan(driver, max_depth=3)
        if not html_laporan:
            time.sleep(10)
            html_laporan = driver.page_source

        soup = BeautifulSoup(html_laporan, 'html.parser')
        semua_tr_soup = [(tr, tr.find_all('td', recursive=False)) for tr in soup.find_all('tr')]
        semua_tr_soup = [(tr, tds) for tr, tds in semua_tr_soup if tds]

        label_kredit_rows = []
        label_umkm_rows = []
        for tr, tds in semua_tr_soup:
            for j, td in enumerate(tds):
                teks = teks_td(td)
                if PATTERN_KREDIT.search(teks):
                    label_kredit_rows.append({'tr': tr, 'tds': tds, 'label_col': j, 'label_text': teks})
                    break
                elif PATTERN_UMKM.search(teks):
                    label_umkm_rows.append({'tr': tr, 'tds': tds, 'label_col': j, 'label_text': teks})
                    break

        hasil_kredit = []
        for info in label_kredit_rows:
            hasil_kredit.append(ambil_jumlah_dari_label_row(info))

        nilai_pihak_terkait_total = hasil_kredit[0] if len(hasil_kredit) >= 1 else "0"
        nilai_pihak_tidak_terkait_total = hasil_kredit[1] if len(hasil_kredit) >= 2 else "0"

        print(f"[HASIL] Kredit – PIHAK TERKAIT     : {nilai_pihak_terkait_total}")
        print(f"[HASIL] Kredit – PIHAK TIDAK TERKAIT : {nilai_pihak_tidak_terkait_total}")

        hasil_umkm = []
        for info in label_umkm_rows:
            hasil_umkm.append(ambil_jumlah_dari_label_row(info))

        nilai_pihak_terkait_umkm = hasil_umkm[0] if len(hasil_kredit) >= 1 else "0"
        nilai_pihak_tidak_terkait_umkm = hasil_umkm[1] if len(hasil_kredit) >= 2 else "0"

        print(f"[HASIL] Kredit UMKM – PIHAK TERKAIT     : {nilai_pihak_terkait_umkm}")
        print(f"[HASIL] Kredit UMKM – PIHAK TIDAK TERKAIT : {nilai_pihak_tidak_terkait_umkm}")

        # Simpan ke memori (List)
        data_rows_excel.append(
            [nomor_urut, '7. Total Kredit – PIHAK TERKAIT', 'J35', nilai_pihak_terkait_total, target_bank, target_tahun, target_bulan,
             nama_laporan])
        nomor_urut += 1
        data_rows_excel.append(
            [nomor_urut, '7. Total Kredit – PIHAK TIDAK TERKAIT', 'J70', nilai_pihak_tidak_terkait_total, target_bank, target_tahun,
             target_bulan, nama_laporan])
        nomor_urut += 1
        data_rows_excel.append(
            [nomor_urut, '7. Kredit UMKM – PIHAK TERKAIT', 'J35', nilai_pihak_terkait_umkm, target_bank, target_tahun, target_bulan,
             nama_laporan])
        nomor_urut += 1
        data_rows_excel.append(
            [nomor_urut, '7. Kredit UMKM – PIHAK TIDAK TERKAIT', 'J70', nilai_pihak_tidak_terkait_umkm, target_bank, target_tahun,
             target_bulan, nama_laporan])
        nomor_urut += 1

# ── 4. EKSPOR KE EXCEL (Setelah Loop Selesai) ─────────────────────────────────
print(f"\n{'=' * 60}\nMenyusun dan Menyimpan Data ke Excel...\n{'=' * 60}")
wb = Workbook()
ws = wb.active
ws.title = "Dataset Panel OJK"


def buat_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)


BORDER = buat_border()
HDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HDR_FILL = PatternFill('solid', start_color='1F4E79')
FILL_KUNING = PatternFill('solid', start_color='FFFF00')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

headers = ['No', 'Keterangan', 'Pos Laporan', 'Nilai Jumlah\n(Juta Rp)', 'Bank', 'Tahun', 'Bulan', 'Laporan']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, CENTER, BORDER

for r_idx, row_data in enumerate(data_rows_excel, 2):
    for c_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = Font(name='Arial', size=10)
        cell.fill = FILL_KUNING
        cell.alignment = CENTER if c_idx in {1, 3, 5, 6, 7} else LEFT
        cell.border = BORDER

col_widths = [5, 35, 14, 18, 44, 8, 12, 28]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

ws.row_dimensions[1].height = 30

nama_file_akhir = f"Hasil_Kredit_{target_bank[:3]}_{target_tahun}_Kuartalan.xlsx"
wb.save(nama_file_akhir)
print(f" SUKSES! Semua data kuartalan berhasil disimpan ke: {nama_file_akhir}")